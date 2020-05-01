import myconnutils
import datetime
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog as fd
from tkinter import ttk
from openpyxl.styles import Alignment
from openpyxl import Workbook
import smtplib
import numpy as np
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg, NavigationToolbar2Tk)
import matplotlib.pyplot as plt

users_all_name = []
all_users_connection = myconnutils.get_connection()
sql = "select employee.fio from diploma.employee;"
try:
    cursor = all_users_connection.cursor()
    cursor.execute(sql)
    for row in cursor:
        for elem in row:
            users_all_name.append(elem)
finally:
    # Закрыть соединение (Close connection).
    all_users_connection.close()


def show_graphic_1():
    line_graph_root = Tk()
    line_graph_root.wm_title('Количество посетителей за месяц')
    line_graph_root["bg"] = "#c9d0a4"

    fig = plt.figure()
    ax = fig.add_subplot(111)
    ax.set(xlim=[0, 31], ylim=[0, 100], title='Количество посетителей за месяц', xlabel='Время, дни',
           ylabel='Количество посетителей, чел')
    ax.plot(
        [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30,
         31],
        [15, 16, 18, 14, 16, 10, 8, 22, 16, 26, 15, 28, 6, 12, 28, 20, 23, 18, 20, 5, 9, 20, 22, 24,
         25, 13, 2, 8, 23, 20, 21],
        color='green', linewidth=2)

    # plt.show()

    canvas = FigureCanvasTkAgg(fig, master=line_graph_root)
    canvas.draw()
    canvas.get_tk_widget().pack(side=TOP, fill=BOTH, expand=1)

    toolbar = NavigationToolbar2Tk(canvas, line_graph_root)
    toolbar.update()
    canvas.get_tk_widget().pack(side=TOP, fill=BOTH, expand=1)

    line_graph_root.mainloop()


def show_diagram_2():
    diagram_2_info = []
    diagram_connection = myconnutils.get_connection()
    diagram_sql = "Select log.point_id,count(log.id_log) from diploma.log group by log.point_id;"
    try:
        diagram_cursor = diagram_connection.cursor()
        diagram_cursor.execute(diagram_sql)
        for diagram_row in diagram_cursor:
            diagram_2_info.append(diagram_row)

    finally:
        # Закрыть соединение (Close connection).
        diagram_connection.close()

    diagram_root = Tk()
    diagram_root.wm_title('Нагрузка на точки входа')
    diagram_root["bg"] = "#c9d0a4"

    nums = []

    for num in diagram_2_info:
        nums.append(num[1])

    labels = 'Главный вход', 'Запасной вход'

    sizes = [nums[0], nums[1]]
    explode = (0, 0.1)  # only "explode" the 2nd slice (i.e. 'Запасной вход')

    fig1, ax1 = plt.subplots()
    ax1.pie(sizes, explode=explode, labels=labels, autopct='%1.1f%%',
            shadow=True, startangle=90)
    ax1.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.

    canvas = FigureCanvasTkAgg(fig1, master=diagram_root)
    canvas.draw()
    canvas.get_tk_widget().pack(side=TOP, fill=BOTH, expand=1)

    toolbar = NavigationToolbar2Tk(canvas, diagram_root)
    toolbar.update()
    canvas.get_tk_widget().pack(side=TOP, fill=BOTH, expand=1)

    diagram_root.mainloop()


def btn_show_log():
    def save_log_as(log_table):
        file_name = fd.asksaveasfilename(filetypes=(("EXCEL files", "*.xlsx"),
                                                    ("TXT files", "*.txt"),
                                                    ("HTML files", "*.html;*.htm"),
                                                    ("All files", "*.*"),), defaultextension='.xlsx')
        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Log_info'

        row_excel = 1
        sheet['A' + str(row_excel)] = 'Событие'
        sheet['B' + str(row_excel)] = 'ФИО'
        sheet['C' + str(row_excel)] = 'UUiD'
        sheet['D' + str(row_excel)] = 'Отдел'
        sheet['E' + str(row_excel)] = 'Точка прохода'
        sheet['F' + str(row_excel)] = 'Время события'

        for i, log in enumerate(log_table):
            sheet['A' + str(i + 2)] = log[0]
            sheet['B' + str(i + 2)] = log[1]
            sheet['C' + str(i + 2)] = log[2]
            sheet['D' + str(i + 2)] = log[3]
            sheet['E' + str(i + 2)] = log[4]
            sheet['F' + str(i + 2)] = log[5]

        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 35
        sheet.column_dimensions['D'].width = 30
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 20

        sheet['A' + str(row_excel)].alignment = Alignment(horizontal='center')
        sheet['B' + str(row_excel)].alignment = Alignment(horizontal='center')
        sheet['C' + str(row_excel)].alignment = Alignment(horizontal='center')
        sheet['D' + str(row_excel)].alignment = Alignment(horizontal='center')
        sheet['E' + str(row_excel)].alignment = Alignment(horizontal='center')
        sheet['F' + str(row_excel)].alignment = Alignment(horizontal='center')

        wb.save(file_name)

    def month_select():
        log_root.destroy()
        sql_log = "Select log.entry_type,employee.fio,employee.UUid,department.name,points.name,log.reg_date " \
                  "from diploma.log inner join diploma.employee on log.employee_id =employee.id_employee " \
                  "inner join department on employee.department_code =department.id_dep " \
                  "inner join points on log.point_id=points.id_point " \
                  "where log.reg_date>CURRENT_TIMESTAMP - INTERVAL 30 DAY;"
        sub_log_root = Tk()
        sub_log_root.title('Журнал событий')
        sub_log_root.geometry('1000x500')
        sub_log_root["bg"] = "#c9d0a4"

        style = ttk.Style(sub_log_root)
        style.theme_use("clam")
        style.configure("Treeview", background='#f0dab5')

        log_table = []
        log_table_connection = myconnutils.get_connection()
        try:
            log_table_cursor = log_table_connection.cursor()
            log_table_cursor.execute(sql_log)
            for log_table_row in log_table_cursor:
                log_table.append(log_table_row)
        finally:
            # Закрыть соединение (Close connection).
            log_table_connection.close()

        log_tree = ttk.Treeview(sub_log_root,
                                columns=('Событие', 'ФИО', 'UUiD', 'Отдел', 'Точка прохода', 'Время события',),
                                height=20, show='headings')
        log_tree.column('Событие', width=100, anchor=CENTER)
        log_tree.column('ФИО', width=200, anchor=CENTER)
        log_tree.column('UUiD', width=225, anchor=CENTER)
        log_tree.column('Отдел', width=150, anchor=CENTER)
        log_tree.column('Точка прохода', width=100, anchor=CENTER)
        log_tree.column('Время события', width=150, anchor=CENTER)

        log_tree.heading('Событие', text='Событие')
        log_tree.heading('ФИО', text='ФИО')
        log_tree.heading('UUiD', text='UUiD')
        log_tree.heading('Отдел', text='Отдел')
        log_tree.heading('Точка прохода', text='Точка прохода')
        log_tree.heading('Время события', text='Время события')

        for log in log_table:
            log_tree.insert('', 'end', values=log)

        log_tree.pack()

        save_log_as_btn = Button(sub_log_root, text='Сохранить', command=lambda: save_log_as(log_table),
                                 font='Times 14', width=25, height=2, bg='#e8ddaf', foreground='#817162')
        save_log_as_btn.place(x=370, y=435)

        sub_log_root.mainloop()

    def year_select():
        log_root.destroy()
        sql_log = "Select log.entry_type,employee.fio,employee.UUid,department.name,points.name,log.reg_date " \
                  "from diploma.log inner join diploma.employee on log.employee_id =employee.id_employee " \
                  "inner join department on employee.department_code =department.id_dep " \
                  "inner join points on log.point_id=points.id_point;"

        sub_log_root = Tk()
        sub_log_root.title('Журнал событий')
        sub_log_root.geometry('1000x500')
        sub_log_root["bg"] = "#c9d0a4"

        style = ttk.Style(sub_log_root)
        style.theme_use("clam")
        style.configure("Treeview", background='#f0dab5')

        log_table = []
        log_table_connection = myconnutils.get_connection()
        try:
            log_table_cursor = log_table_connection.cursor()
            log_table_cursor.execute(sql_log)
            for log_table_row in log_table_cursor:
                log_table.append(log_table_row)
        finally:
            # Закрыть соединение (Close connection).
            log_table_connection.close()

        log_tree = ttk.Treeview(sub_log_root,
                                columns=('Событие', 'ФИО', 'UUiD', 'Отдел', 'Точка прохода', 'Время события',),
                                height=20, show='headings')
        log_tree.column('Событие', width=100, anchor=CENTER)
        log_tree.column('ФИО', width=200, anchor=CENTER)
        log_tree.column('UUiD', width=225, anchor=CENTER)
        log_tree.column('Отдел', width=150, anchor=CENTER)
        log_tree.column('Точка прохода', width=100, anchor=CENTER)
        log_tree.column('Время события', width=150, anchor=CENTER)

        log_tree.heading('Событие', text='Событие')
        log_tree.heading('ФИО', text='ФИО')
        log_tree.heading('UUiD', text='UUiD')
        log_tree.heading('Отдел', text='Отдел')
        log_tree.heading('Точка прохода', text='Точка прохода')
        log_tree.heading('Время события', text='Время события')

        for log in log_table:
            log_tree.insert('', 'end', values=log)

        log_tree.pack()

        save_log_as_btn = Button(sub_log_root, text='Сохранить', command=lambda: save_log_as(log_table),
                                 font='Times 14', width=25, height=2, bg='#e8ddaf', foreground='#817162')
        save_log_as_btn.place(x=370, y=435)

        sub_log_root.mainloop()

    log_root = Tk()
    log_root.title('Журнал событий')
    log_root.geometry('300x200')
    log_root["bg"] = "#c9d0a4"

    log_label = Label(log_root, text='Выберите период', font='Times 14', width=35, height=2, bg="#c9d0a4")
    log_label.pack()

    btn_month = Button(log_root, text='За месяц', command=month_select, font='Times 14', width=25, height=2,
                       bg='#e8ddaf', foreground='#817162')
    btn_year = Button(log_root, text='За все время', command=year_select, font='Times 14', width=25, height=2,
                      bg='#e8ddaf', foreground='#817162')
    btn_month.pack()
    btn_year.place(x=20, y=120)

    log_root.mainloop()


def show_all_info():
    def save_as():
        file_name = fd.asksaveasfilename(filetypes=(("EXCEL files", "*.xlsx"),
                                                    ("TXT files", "*.txt"),
                                                    ("HTML files", "*.html;*.htm"),
                                                    ("All files", "*.*"),), defaultextension='.xlsx')
        wb = Workbook()
        sheet = wb.active
        sheet.title = 'User_info'

        row_excel = 1
        sheet['A' + str(row_excel)] = 'ФИО'
        sheet['B' + str(row_excel)] = 'Телефон'
        sheet['C' + str(row_excel)] = 'Email'
        sheet['D' + str(row_excel)] = 'Дата рождения'
        sheet['E' + str(row_excel)] = 'UUiD'
        sheet['F' + str(row_excel)] = 'Статус нахождения на объекте в данный момент'
        sheet['G' + str(row_excel)] = 'Коэффициент пунктуальности'
        sheet['H' + str(row_excel)] = 'Отдел'
        sheet['I' + str(row_excel)] = 'Базовая заработная плата по отделу'
        sheet['J' + str(row_excel)] = 'Должность'
        sheet['K' + str(row_excel)] = 'Надбавка за должность'
        sheet['L' + str(row_excel)] = 'Дата принятия на работу'

        for i, user in enumerate(all_users_info):
            sheet['A' + str(i + 2)] = user[0]
            sheet['B' + str(i + 2)] = user[1]
            sheet['C' + str(i + 2)] = user[2]
            sheet['D' + str(i + 2)] = user[3]
            sheet['E' + str(i + 2)] = user[4]
            sheet['F' + str(i + 2)] = user[5]
            sheet['G' + str(i + 2)] = user[6]
            sheet['H' + str(i + 2)] = user[7]
            sheet['I' + str(i + 2)] = user[8]
            sheet['J' + str(i + 2)] = user[9]
            sheet['K' + str(i + 2)] = user[10]
            sheet['L' + str(i + 2)] = user[11]

        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 18
        sheet.column_dimensions['C'].width = 25
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 36
        sheet.column_dimensions['F'].width = 45
        sheet.column_dimensions['G'].width = 28
        sheet.column_dimensions['H'].width = 20
        sheet.column_dimensions['I'].width = 36
        sheet.column_dimensions['J'].width = 20
        sheet.column_dimensions['K'].width = 25
        sheet.column_dimensions['L'].width = 25

        sheet['A' + str(row_excel)].alignment = Alignment(horizontal='center')
        sheet['B' + str(row_excel)].alignment = Alignment(horizontal='center')
        sheet['C' + str(row_excel)].alignment = Alignment(horizontal='center')
        sheet['D' + str(row_excel)].alignment = Alignment(horizontal='center')
        sheet['E' + str(row_excel)].alignment = Alignment(horizontal='center')
        sheet['F' + str(row_excel)].alignment = Alignment(horizontal='center')
        sheet['G' + str(row_excel)].alignment = Alignment(horizontal='center')
        sheet['H' + str(row_excel)].alignment = Alignment(horizontal='center')
        sheet['I' + str(row_excel)].alignment = Alignment(horizontal='center')
        sheet['J' + str(row_excel)].alignment = Alignment(horizontal='center')
        sheet['K' + str(row_excel)].alignment = Alignment(horizontal='center')
        sheet['L' + str(row_excel)].alignment = Alignment(horizontal='center')

        wb.save(file_name)

    all_users_info = []
    all_users_info_connection = myconnutils.get_connection()
    sql_all_info = "Select employee.fio, employee.phone, employee.email, employee.date_of_birth," \
                   " employee.uuid, employee.entry_check, employee.late_odd ," \
                   "department.name, department.salary_value, odds.name, odds.odd_value, employee.reg_date" \
                   " from employee inner join department on employee.department_code = department.id_dep " \
                   "inner join odds on employee.odd_code=odds.id_odd ;"
    try:
        all_user_info_cursor = all_users_info_connection.cursor()
        all_user_info_cursor.execute(sql_all_info)
        for all_users_info_row in all_user_info_cursor:
            all_users_info.append(all_users_info_row)
    finally:
        # Закрыть соединение (Close connection).
        all_users_info_connection.close()

    all_root = Tk()
    all_root.title("Все пользователи")
    all_root.geometry("1650x400")
    all_root["bg"] = "#c9d0a4"

    style = ttk.Style(all_root)
    style.theme_use("clam")
    style.configure("Treeview", background='#f0dab5')

    tree = ttk.Treeview(all_root, columns=('ФИО', 'Телефон', 'Email', 'Дата рождения', 'UUiD',
                                           'Статус',
                                           'Коэффициент пунктуальности', 'Отдел', 'Базовая з/п'
                                           , 'Должность', 'Надбавка за должность', 'Дата принятия на работу',),
                        height=15, show='headings')

    tree.column('ФИО', width=200, anchor=CENTER)
    tree.column('Телефон', width=100, anchor=CENTER)
    tree.column('Email', width=170, anchor=CENTER)
    tree.column('Дата рождения', width=90, anchor=CENTER)
    tree.column('UUiD', width=220, anchor=CENTER)
    tree.column('Статус', width=50, anchor=CENTER)
    tree.column('Коэффициент пунктуальности', width=190, anchor=CENTER)
    tree.column('Отдел', width=150, anchor=CENTER)
    tree.column('Базовая з/п', width=90, anchor=CENTER)
    tree.column('Должность', width=130, anchor=CENTER)
    tree.column('Надбавка за должность', width=60, anchor=CENTER)
    tree.column('Дата принятия на работу', width=150, anchor=CENTER)

    tree.heading('ФИО', text='ФИО')
    tree.heading('Телефон', text='Телефон')
    tree.heading('Email', text='Email')
    tree.heading('Дата рождения', text='Дата рождения')
    tree.heading('UUiD', text='UUiD')
    tree.heading('Статус', text='Статус')
    tree.heading('Коэффициент пунктуальности', text='Коэффициент пунктуальности')
    tree.heading('Отдел', text='Отдел')
    tree.heading('Базовая з/п', text='Базовая з/п')
    tree.heading('Должность', text='Должность')
    tree.heading('Надбавка за должность', text='Надбавка за должность')
    tree.heading('Дата принятия на работу', text='Дата принятия на работу')

    for employee_one in all_users_info:
        tree.insert('', 'end', values=employee_one)

    tree.pack()
    save_as_btn = Button(all_root, text='Сохранить', command=save_as, font='Times 14', width=35, height=2, bg='#e8ddaf',
                         foreground='#817162')
    save_as_btn.place(x=620, y=335)

    all_root.mainloop()


def show_info():
    def user_request():
        if show_info_entry.get() == '' or show_info_entry.get() not in users_all_name:

            user_info_error_root = Tk()
            user_info_error_root.title('Ошибка пользователя')
            user_info_error_root.geometry('400x135')
            user_info_error_root["bg"] = "#c9d0a4"

            find_info_label = Label(user_info_error_root, text='Пользователя с данным именем не существует!',
                                    font='Times 14', width=40, height=2, bg="#c9d0a4")
            find_info_btn = Button(user_info_error_root, text='OK', command=user_info_error_root.destroy,
                                   font='Times 14', width=30, height=2, bg='#e8ddaf', foreground='#817162')
            find_info_label.pack()
            find_info_btn.pack()

            user_info_error_root.mainloop()

        else:
            current_employee_info = ''
            db_info_connection = myconnutils.get_connection()
            sql_find_info = "Select employee.fio, employee.phone, employee.email, employee.date_of_birth," \
                            " employee.uuid, employee.entry_check, employee.reg_date, employee.late_odd ," \
                            " department.name, department.salary_value, odds.name, odds.odd_value" \
                            " from employee inner join department on employee.department_code = department.id_dep " \
                            "inner join odds on employee.odd_code=odds.id_odd  " \
                            "where employee.fio ='{0}';".format(show_info_entry.get())
            try:
                db_info_cursor = db_info_connection.cursor()
                db_info_cursor.execute(sql_find_info)
                for db_info_row in db_info_cursor:
                    current_employee_info = db_info_row
            finally:
                # Закрыть соединение (Close connection).
                db_info_connection.close()

            #######################
            # User graphic
            #######################
            def show_graphic_1():
                graph_list = []
                line_graph_connection = myconnutils.get_connection()
                line_graph_sql = "select log.id_log,log.reg_date, log.entry_type from diploma.log " \
                                 "inner join diploma.employee on log.employee_id = employee.id_employee" \
                                 " where employee.fio='{0}';".format(current_employee_info[0])
                try:
                    line_graph_cursor = line_graph_connection.cursor()
                    line_graph_cursor.execute(line_graph_sql)
                    for line_graph_row in line_graph_cursor:
                        graph_list.append(line_graph_row)
                finally:
                    # Закрыть соединение (Close connection).
                    line_graph_connection.close()

                line_graph_root = Tk()
                line_graph_root.wm_title('Отработка часов за месяц')
                line_graph_root["bg"] = "#c9d0a4"

                x = np.arange(1, 31)
                y = np.random.randint(6, 10, size=30)

                fig, ax = plt.subplots()

                ax.bar(x, y)
                ax.set(title='Отработка часов за месяц', xlabel='Время, дни',
                       ylabel='Время, ч')

                ax.set_facecolor('seashell')
                fig.set_facecolor('floralwhite')
                fig.set_figwidth(12)  # ширина Figure
                fig.set_figheight(6)  # высота Figure

                plt.show()

                canvas = FigureCanvasTkAgg(fig, master=line_graph_root)
                canvas.draw()
                canvas.get_tk_widget().pack(side=TOP, fill=BOTH, expand=1)

                toolbar = NavigationToolbar2Tk(canvas, line_graph_root)
                toolbar.update()
                canvas.get_tk_widget().pack(side=TOP, fill=BOTH, expand=1)

                line_graph_root.mainloop()

            #######################
            # User graphic
            #######################

            def save_file():
                file_name = fd.asksaveasfilename(filetypes=(("EXCEL files", "*.xlsx"),
                                                            ("TXT files", "*.txt"),
                                                            ("HTML files", "*.html;*.htm"),
                                                            ("All files", "*.*"),), defaultextension='.xlsx')
                wb = Workbook()
                sheet = wb.active
                sheet.title = 'User_info'

                row_excel = 1
                sheet['A' + str(row_excel)] = 'ФИО'
                sheet['B' + str(row_excel)] = 'Телефон'
                sheet['C' + str(row_excel)] = 'Email'
                sheet['D' + str(row_excel)] = 'Дата рождения'
                sheet['E' + str(row_excel)] = 'UUiD'
                sheet['F' + str(row_excel)] = 'Статус нахождения на объекте в данный момент'
                sheet['G' + str(row_excel)] = 'Коэффициент пунктуальности'
                sheet['H' + str(row_excel)] = 'Отдел'
                sheet['I' + str(row_excel)] = 'Базовая заработная плата по отделу'
                sheet['J' + str(row_excel)] = 'Должность'
                sheet['K' + str(row_excel)] = 'Надбавка за должность'
                sheet['L' + str(row_excel)] = 'Дата принятия на работу'

                sheet['A' + str(2)] = current_employee_info[0]
                sheet['B' + str(2)] = current_employee_info[1]
                sheet['C' + str(2)] = current_employee_info[2]
                sheet['D' + str(2)] = current_employee_info[3]
                sheet['E' + str(2)] = current_employee_info[4]

                if current_employee_info[5] == 0:
                    sheet['F' + str(2)] = 'Присутствует'
                else:
                    sheet['F' + str(2)] = 'Отсутствует'

                sheet['G' + str(2)] = current_employee_info[7]
                sheet['H' + str(2)] = current_employee_info[8]
                sheet['I' + str(2)] = current_employee_info[9]
                sheet['J' + str(2)] = current_employee_info[10]
                sheet['K' + str(2)] = current_employee_info[11]
                sheet['L' + str(2)] = current_employee_info[6]

                sheet.column_dimensions['A'].width = 30
                sheet.column_dimensions['B'].width = 18
                sheet.column_dimensions['C'].width = 25
                sheet.column_dimensions['D'].width = 15
                sheet.column_dimensions['E'].width = 36
                sheet.column_dimensions['F'].width = 45
                sheet.column_dimensions['G'].width = 28
                sheet.column_dimensions['H'].width = 20
                sheet.column_dimensions['I'].width = 36
                sheet.column_dimensions['J'].width = 20
                sheet.column_dimensions['K'].width = 25
                sheet.column_dimensions['L'].width = 25

                sheet['A' + str(row_excel)].alignment = Alignment(horizontal='center')
                sheet['B' + str(row_excel)].alignment = Alignment(horizontal='center')
                sheet['C' + str(row_excel)].alignment = Alignment(horizontal='center')
                sheet['D' + str(row_excel)].alignment = Alignment(horizontal='center')
                sheet['E' + str(row_excel)].alignment = Alignment(horizontal='center')
                sheet['F' + str(row_excel)].alignment = Alignment(horizontal='center')
                sheet['G' + str(row_excel)].alignment = Alignment(horizontal='center')
                sheet['H' + str(row_excel)].alignment = Alignment(horizontal='center')
                sheet['I' + str(row_excel)].alignment = Alignment(horizontal='center')
                sheet['J' + str(row_excel)].alignment = Alignment(horizontal='center')
                sheet['K' + str(row_excel)].alignment = Alignment(horizontal='center')
                sheet['L' + str(row_excel)].alignment = Alignment(horizontal='center')

                sheet['A' + str(2)].alignment = Alignment(horizontal='center')
                sheet['B' + str(2)].alignment = Alignment(horizontal='center')
                sheet['C' + str(2)].alignment = Alignment(horizontal='center')
                sheet['D' + str(2)].alignment = Alignment(horizontal='center')
                sheet['E' + str(2)].alignment = Alignment(horizontal='center')
                sheet['F' + str(2)].alignment = Alignment(horizontal='center')
                sheet['G' + str(2)].alignment = Alignment(horizontal='center')
                sheet['H' + str(2)].alignment = Alignment(horizontal='center')
                sheet['I' + str(2)].alignment = Alignment(horizontal='center')
                sheet['J' + str(2)].alignment = Alignment(horizontal='center')
                sheet['K' + str(2)].alignment = Alignment(horizontal='center')
                sheet['L' + str(2)].alignment = Alignment(horizontal='center')

                wb.save(file_name)

            show_info_label.forget()
            show_info_entry.forget()
            show_info_btn.forget()

            info_user_fio_label = Label(user_info_root, text='ФИО', font='Times 14', width=40, height=2, bg="#c9d0a4")
            info_user_fio_label.pack()

            info_user_fio_entry = Entry(user_info_root, justify=CENTER, width=40, font='Times 14', bg="#c9d0a4")
            info_user_fio_entry.insert(0, current_employee_info[0])
            info_user_fio_entry.config(state=DISABLED)
            info_user_fio_entry.pack()

            info_user_phone_label = Label(user_info_root, text='Телефон', font='Times 14', width=40, height=2,
                                          bg="#c9d0a4")
            info_user_phone_label.pack()

            info_user_phone_entry = Entry(user_info_root, justify=CENTER, font='Times 14', width=40, bg="#c9d0a4")
            info_user_phone_entry.insert(0, current_employee_info[1])
            info_user_phone_entry.config(state=DISABLED)
            info_user_phone_entry.pack()

            info_user_email_label = Label(user_info_root, text='Email', font='Times 14', width=40, height=2,
                                          bg="#c9d0a4")
            info_user_email_label.pack()

            info_user_email_entry = Entry(user_info_root, justify=CENTER, font='Times 14', width=40, bg="#c9d0a4")
            info_user_email_entry.insert(0, current_employee_info[2])
            info_user_email_entry.config(state=DISABLED)
            info_user_email_entry.pack()

            info_user_date_of_birth_label = Label(user_info_root, text='Дата рождения', font='Times 14', width=40,
                                                  height=2, bg="#c9d0a4")
            info_user_date_of_birth_label.pack()

            info_user_date_of_birth_entry = Entry(user_info_root, justify=CENTER, font='Times 14', width=40,
                                                  bg="#c9d0a4")
            info_user_date_of_birth_entry.insert(0, current_employee_info[3])
            info_user_date_of_birth_entry.config(state=DISABLED)
            info_user_date_of_birth_entry.pack()

            info_user_uuid_label = Label(user_info_root, text='UUiD', font='Times 14', width=40, height=2, bg="#c9d0a4")
            info_user_uuid_label.pack()

            info_user_uuid_entry = Entry(user_info_root, justify=CENTER, font='Times 14', width=40, bg="#c9d0a4")
            info_user_uuid_entry.insert(0, current_employee_info[4])
            info_user_uuid_entry.config(state=DISABLED)
            info_user_uuid_entry.pack()

            info_user_entry_check_label = Label(user_info_root, text='Статус нахождения на объекте в данный момент',
                                                font='Times 14', width=40, height=2, bg="#c9d0a4")
            info_user_entry_check_label.pack()

            info_user_entry_check_entry = Entry(user_info_root, justify=CENTER, font='Times 14', width=40, bg="#c9d0a4")
            if current_employee_info[5] == 0:
                info_user_entry_check_entry.insert(0, 'Отсутствует')
            else:
                info_user_entry_check_entry.insert(0, 'На рабочем месте')
            info_user_entry_check_entry.config(state=DISABLED)
            info_user_entry_check_entry.pack()

            info_user_late_odd_label = Label(user_info_root, text='Коэффициент пунктуальности', font='Times 14',
                                             width=40, height=2, bg="#c9d0a4")
            info_user_late_odd_label.pack()

            info_user_late_odd_entry = Entry(user_info_root, justify=CENTER, font='Times 14', width=40, bg="#c9d0a4")
            info_user_late_odd_entry.insert(0, current_employee_info[7])
            info_user_late_odd_entry.config(state=DISABLED)
            info_user_late_odd_entry.pack()

            info_user_department_name_label = Label(user_info_root, text='Отдел', font='Times 14', width=40, height=2,
                                                    bg="#c9d0a4")
            info_user_department_name_label.pack()

            info_user_department_name_entry = Entry(user_info_root, justify=CENTER, font='Times 14', width=40,
                                                    bg="#c9d0a4")
            info_user_department_name_entry.insert(0, current_employee_info[8])
            info_user_department_name_entry.config(state=DISABLED)
            info_user_department_name_entry.pack()

            info_user_department_salary_label = Label(user_info_root, text='Базовая заработная плата по отделу',
                                                      font='Times 14', width=40, height=2, bg="#c9d0a4")
            info_user_department_salary_label.pack()

            info_user_department_salary_entry = Entry(user_info_root, justify=CENTER, font='Times 14', width=40,
                                                      bg="#c9d0a4")
            info_user_department_salary_entry.insert(0, current_employee_info[9])
            info_user_department_salary_entry.config(state=DISABLED)
            info_user_department_salary_entry.pack()

            info_user_odd_name_label = Label(user_info_root, text='Должность', font='Times 14', width=40, height=2,
                                             bg="#c9d0a4")
            info_user_odd_name_label.pack()

            info_user_odd_name_entry = Entry(user_info_root, justify=CENTER, font='Times 14', width=40, bg="#c9d0a4")
            info_user_odd_name_entry.insert(0, current_employee_info[10])
            info_user_odd_name_entry.config(state=DISABLED)
            info_user_odd_name_entry.pack()

            info_user_odd_value_label = Label(user_info_root, text='Надбавка за должность', font='Times 14', width=40,
                                              height=2, bg="#c9d0a4")
            info_user_odd_value_label.pack()

            info_user_odd_value_entry = Entry(user_info_root, justify=CENTER, font='Times 14', width=40, bg="#c9d0a4")
            info_user_odd_value_entry.insert(0, current_employee_info[11])
            info_user_odd_value_entry.config(state=DISABLED)
            info_user_odd_value_entry.pack()

            info_user_reg_date_label = Label(user_info_root, text='Дата принятия на работу', font='Times 14', width=40,
                                             height=2, bg="#c9d0a4")
            info_user_reg_date_label.pack()

            info_user_reg_date_entry = Entry(user_info_root, justify=CENTER, font='Times 14', width=40, bg="#c9d0a4")
            info_user_reg_date_entry.insert(0, current_employee_info[6])
            info_user_reg_date_entry.config(state=DISABLED)
            info_user_reg_date_entry.pack()

            info_user_save_btn = Button(user_info_root, text='Сохранить', command=save_file, font='Times 14', width=35,
                                        height=1, bg='#e8ddaf', foreground='#817162')
            info_user_save_btn.place(x=70, y=890)

            info_graphic_user_btn = Button(user_info_root, text='График посещения', command=show_graphic_1,
                                           font='Times 14', width=35, height=1, bg='#e8ddaf', foreground='#817162')
            info_graphic_user_btn.place(x=70, y=940)

    user_info_root = Tk()
    user_info_root.title('Информация о пользователе')
    user_info_root.geometry('500x1000')
    user_info_root["bg"] = "#c9d0a4"

    show_info_label = Label(user_info_root, text='Введите ФИО пользователя', font='Times 14', width=40, height=2,
                            bg="#c9d0a4")
    show_info_entry = Entry(user_info_root, font='Times 14', width=40, bg="#c9d0a4")
    show_info_btn = Button(user_info_root, text='Поиск пользователя', command=user_request, font='Times 14', width=30,
                           height=2, bg='#e8ddaf', foreground='#817162')

    show_info_label.pack()
    show_info_entry.pack()
    show_info_btn.place(x=100, y=100)

    user_info_root.mainloop()


def email_send_back(door):
    content = 'Несанкционированная попытка доступа но объект! Просьба проверить точку доступа № {0}'.format(
        door).encode('utf-8')
    mail = smtplib.SMTP('smtp.gmail.com', 587)
    mail.ehlo()
    mail.starttls()
    mail.login('shopmanage7@gmail.com', 'M2t8zUmPQg')
    mail.sendmail('shopmanage7@gmail.com', 'bardiervadim97@gmail.com', content)
    mail.close()


def user_update():
    def update_userprofile():

        def update_confirm():
            up_fio = up_fio_entry.get()
            up_phone = up_phone_entry.get()
            up_email = up_email_entry.get()
            up_date_of_birth = up_date_of_birth_entry.get()
            up_uuid = up_uuid_entry.get()
            up_department_choice = up_departments_listbox.curselection()[0] + 1  # Возвращает индекс+1
            up_odd_choice = up_odds_listbox.curselection()[0] + 1  # Возвращает индекс+1

            connec = myconnutils.get_connection()
            sql_3 = "UPDATE diploma.employee SET fio='{0}',phone='{1}',email='{2}',date_of_birth='{3}',UUid='{4}'," \
                    "department_code={5},odd_code={6} WHERE employee.fio='{7}';".format(up_fio, up_phone, up_email,
                                                                                        up_date_of_birth, up_uuid,
                                                                                        up_department_choice,
                                                                                        up_odd_choice, user_fio)
            try:
                curs = connec.cursor()
                curs.execute(sql_3)
                connec.commit()
                messagebox.showinfo("Успех!", message='Пользователь успешно обновлен!')
            except:
                connec.rollback()
            finally:

                connec.close()
            sub_sub_root.destroy()

        user_fio = update_entry.get()

        if user_fio == '' or user_fio not in users_all_name:
            sub_sub_root = Tk()
            sub_sub_root.title('Ошибка пользователя')
            sub_sub_root.geometry('600x120')
            sub_sub_root["bg"] = "#c9d0a4"

            up_label = Label(sub_sub_root, text='Пользователя с именем {0} не существует!'.format(user_fio),
                             font='Times 14', width=60, height=2, bg="#c9d0a4")
            up_btn = Button(sub_sub_root, text='OK', command=sub_sub_root.destroy, font='Times 14', width=30, height=2,
                            bg='#e8ddaf', foreground='#817162')
            up_label.pack()
            up_btn.pack()

            sub_sub_root.mainloop()
        else:
            sub_sub_root = Tk()
            sub_sub_root.title('Изменение пользователя: {}'.format(user_fio))
            sub_sub_root.geometry('500x900')
            employee_info = ''
            sub_sub_root["bg"] = "#c9d0a4"

            connection = myconnutils.get_connection()
            sql = "Select * from diploma.employee where employee.fio='{}';".format(update_entry.get())
            sql_1 = "Select * from diploma.department;"
            sql_2 = "Select * from diploma.odds;"
            departments_list = []
            odds_list = []

            try:
                cursor = connection.cursor()
                cursor.execute(sql)
                for row in cursor:
                    employee_info = row
                cursor = connection.cursor()
                cursor.execute(sql_1)
                for row in cursor:
                    departments_list.append(row[1])
                cursor.execute(sql_2)
                for row in cursor:
                    odds_list.append(row[1])
            finally:
                # Закрыть соединение (Close connection).
                connection.close()

            up_fio_label = Label(sub_sub_root, text='Новое ФИО', font='Times 14', width=40, height=2, bg="#c9d0a4")
            up_fio_entry = Entry(sub_sub_root, justify=CENTER, font='Times 14', width=40, bg="#c9d0a4")
            up_fio_entry.insert(0, employee_info[1])

            up_phone_label = Label(sub_sub_root, text='Новый Телефон', font='Times 14', width=40, height=2,
                                   bg="#c9d0a4")
            up_phone_entry = Entry(sub_sub_root, justify=CENTER, font='Times 14', width=40, bg="#c9d0a4")
            up_phone_entry.insert(0, employee_info[2])

            up_email_label = Label(sub_sub_root, text='Новый Email', font='Times 14', width=40, height=2, bg="#c9d0a4")
            up_email_entry = Entry(sub_sub_root, justify=CENTER, font='Times 14', width=40, bg="#c9d0a4")
            up_email_entry.insert(0, employee_info[3])

            up_date_of_birth_label = Label(sub_sub_root, text='Новая Дата рождения', font='Times 14', width=40,
                                           height=2, bg="#c9d0a4")
            up_date_of_birth_entry = Entry(sub_sub_root, justify=CENTER, font='Times 14', width=40, bg="#c9d0a4")
            up_date_of_birth_entry.insert(0, employee_info[4])

            up_uuid_label = Label(sub_sub_root, text='UUiD', font='Times 14', width=40, height=2, bg="#c9d0a4")
            up_uuid_entry = Entry(sub_sub_root, justify=CENTER, font='Times 14', width=40, bg="#c9d0a4")
            up_uuid_entry.insert(0, employee_info[5])

            up_department_label = Label(sub_sub_root, text='Отдел', font='Times 14', width=40, height=2, bg="#c9d0a4")
            up_departments_listbox = Listbox(sub_sub_root, width=40, background='#f0dab5', selectbackground='#ab9379',
                                             exportselection=0,
                                             font='Times 14', height=11)
            scrollbar = Scrollbar(sub_sub_root, command=up_departments_listbox.yview)
            scrollbar.pack(side=RIGHT, fill=Y)
            up_departments_listbox.config(yscrollcommand=scrollbar.set)
            for department in departments_list:
                up_departments_listbox.insert(END, department)

            up_odd_label = Label(sub_sub_root, text='Коэффициент должности', font='Times 14', width=40, height=2,
                                 bg="#c9d0a4")
            up_odds_listbox = Listbox(sub_sub_root, width=40, background='#f0dab5', selectbackground='#ab9379',
                                      font='Times 14', height=4)
            for odd in odds_list:
                up_odds_listbox.insert(END, odd)

            up_btn_confirm = Button(sub_sub_root, text='Изменить пользователя', command=update_confirm, font='Times 14',
                                    width=35, height=2, bg='#e8ddaf', foreground='#817162')

            up_fio_label.pack()
            up_fio_entry.pack()

            up_phone_label.pack()
            up_phone_entry.pack()

            up_email_label.pack()
            up_email_entry.pack()

            up_date_of_birth_label.pack()
            up_date_of_birth_entry.pack()

            up_uuid_label.pack()
            up_uuid_entry.pack()

            up_department_label.pack()
            up_departments_listbox.pack()
            scrollbar.config(command=up_departments_listbox.yview)

            up_odd_label.pack()
            up_odds_listbox.pack()

            up_btn_confirm.place(x=60, y=820)

            sub_root.destroy()
            sub_sub_root.mainloop()

    sub_root = Tk()
    sub_root.title('Изменение существующего пользователя')
    sub_root.geometry('250x150')
    sub_root["bg"] = "#c9d0a4"

    update_label = Label(sub_root, text='Введите ФИО пользователя', font='Times 14', width=40, height=2, bg="#c9d0a4")
    update_entry = Entry(sub_root, width=40, font='Times 14', bg="#c9d0a4")
    update_btn = Button(sub_root, text='Поиск пользователя', command=update_userprofile, font='Times 14', height=2,
                        bg='#e8ddaf', foreground='#817162')

    update_label.pack()
    update_entry.pack()
    update_btn.place(x=40, y=80)

    sub_root.mainloop()


def user_create():
    connection = myconnutils.get_connection()
    sql_1 = "Select * from diploma.department;"
    sql_2 = "Select * from diploma.odds;"
    departments_list = []
    odds_list = []
    try:
        cursor = connection.cursor()
        cursor.execute(sql_1)
        for row in cursor:
            departments_list.append(row[1])
        cursor.execute(sql_2)
        for row in cursor:
            odds_list.append(row[1])
    finally:
        connection.close()

    def accept_create():
        fio = entry_fio.get()
        phone = entry_phone.get()
        email = entry_email.get()
        date_of_birth = entry_date_of_birth.get()
        uuid = entry_uuid.get()
        department_choice = departments_listbox.curselection()[0] + 1  # Возвращает индекс+1
        odd_choice = odds_listbox.curselection()[0] + 1  # Возвращает индекс+1
        conn = myconnutils.get_connection()
        sql = "Insert Employee(fio,phone,email,date_of_birth,UUid,department_code,odd_code,reg_date) values ('{0}'," \
              "'{1}','{2}','{3}','{4}',{5},{6},now());".format(fio, phone, email, date_of_birth, uuid,
                                                               department_choice + 1, odd_choice + 1)
        try:
            cur = conn.cursor()
            cur.execute(sql)
            conn.commit()
            messagebox.showinfo("Успех!", message='Пользователь успешно создан!')
        except:
            conn.rollback()
        finally:
            conn.close()
        sub_root.destroy()

    sub_root = Tk()
    sub_root.title('Создание новго пользователя')
    sub_root.geometry('500x900')
    sub_root["bg"] = "#c9d0a4"

    label_fio = Label(sub_root, text='Фамилия имя отчество', font='Times 14', width=40, height=2, bg="#c9d0a4")
    entry_fio = Entry(sub_root, justify=CENTER, font='Times 14', width=40, bg="#c9d0a4")

    label_phone = Label(sub_root, text='Телефон', font='Times 14', width=40, height=2, bg="#c9d0a4")
    entry_phone = Entry(sub_root, justify=CENTER, font='Times 14', width=40, bg="#c9d0a4")

    label_email = Label(sub_root, text='Email', font='Times 14', width=40, height=2, bg="#c9d0a4")
    entry_email = Entry(sub_root, justify=CENTER, font='Times 14', width=40, bg="#c9d0a4")

    label_date_of_birth = Label(sub_root, text='Дата рождения', font='Times 14', width=40, height=2, bg="#c9d0a4")
    entry_date_of_birth = Entry(sub_root, justify=CENTER, font='Times 14', width=40, bg="#c9d0a4")

    label_uuid = Label(sub_root, text='UUiD', font='Times 14', width=40, height=2, bg="#c9d0a4")
    entry_uuid = Entry(sub_root, justify=CENTER, font='Times 14', width=40, bg="#c9d0a4")

    label_department = Label(sub_root, text='Отдел', font='Times 14', width=40, height=2, bg="#c9d0a4")
    departments_listbox = Listbox(sub_root, width=40, selectbackground='#ab9379', exportselection=0, font='Times 14',
                                  height=11, background='#f0dab5')
    scrollbar = Scrollbar(sub_root, command=departments_listbox.yview)
    scrollbar.pack(side=RIGHT, fill=Y)
    departments_listbox.config(yscrollcommand=scrollbar.set)
    for department in departments_list:
        departments_listbox.insert(END, department)

    label_odd = Label(sub_root, text='Коэффициент должности', font='Times 14', width=40, height=2, bg="#c9d0a4")
    odds_listbox = Listbox(sub_root, width=40, font='Times 14', height=4, background='#f0dab5',
                           selectbackground='#ab9379')
    for odd in odds_list:
        odds_listbox.insert(END, odd)

    btn_accept_create = Button(sub_root, text='Создать', command=accept_create, font='Times 14', width=35, height=2,
                               bg='#e8ddaf', foreground='#817162')

    label_fio.pack()
    entry_fio.pack()

    label_phone.pack()
    entry_phone.pack()

    label_email.pack()
    entry_email.pack()

    label_date_of_birth.pack()
    entry_date_of_birth.pack()

    label_uuid.pack()
    entry_uuid.pack()

    label_department.pack()
    departments_listbox.pack()
    scrollbar.config(command=departments_listbox.yview)

    label_odd.pack()
    odds_listbox.pack()

    btn_accept_create.place(x=60, y=820)

    sub_root.mainloop()


def user_auth():
    sub_auth_root = Tk()
    sub_auth_root.title('Аутенитификация пользователя')
    sub_auth_root.geometry('500x500')
    sub_auth_root["bg"] = "#c9d0a4"

    users = []
    points = []
    uuids = []
    users_uuids = []

    connection = myconnutils.get_connection()
    sql_1 = "Select employee.fio,employee.UUid from employee order by employee.fio;"
    sql_2 = "Select points.name from points;"
    try:
        cursor = connection.cursor()
        cursor.execute(sql_1)
        for row in cursor:
            users_uuids.append(row)
            users.append(row[0])
            uuids.append(row[1])
        users.append('-----')
        uuids.append('af89ac97-b14b-4660-a382-93b26ddf877d')
        uuids.append('76488e31-5707-4603-b571-7f0e1615f87c')
        uuids.append('b5f711e3-153b-4df4-b012-ab5ca2bf2f04')

        cursor.execute(sql_2)
        for row in cursor:
            points.append(row[0])
    finally:
        # Закрыть соединение (Close connection).
        connection.close()

    variable = StringVar(sub_auth_root)
    variable.set(users[-1])  # default value

    fio_label = Label(sub_auth_root, text='Выберите пропуск который хотите приложить', font='Times 14', width=40,
                      height=2, bg="#c9d0a4")
    fio_label.pack()

    def setter_auth(selection):
        variable.set(selection)

    w = OptionMenu(*(sub_auth_root, variable) + tuple(users, ), command=setter_auth)
    w.pack()

    w.config(bg='#c9d0a4', font='Times 14', activebackground='#c9d0a4')

    def submit():
        btn_submit.pack_forget()
        w.pack_forget()

        sub_entry = Entry(sub_auth_root, justify=CENTER, font='Times 14', width=40, bg="#c9d0a4")
        sub_entry.insert(0, variable.get())
        sub_entry.config(state=DISABLED)
        sub_entry.pack()

        label_uuid_auth = Label(sub_auth_root, text='UUiD', font='Times 14', width=40, bg="#c9d0a4")
        label_uuid_auth.pack()
        if variable.get() == '-----':
            entry_uuid_auth = Entry(sub_auth_root, justify=CENTER, font='Times 14', width=40, bg="#c9d0a4")
            entry_uuid_auth.insert(0, 'Введите UUiD')
            entry_uuid_auth.pack()
        else:
            entry_uuid_auth = Entry(sub_auth_root, justify=CENTER, font='Times 14', width=40, bg="#c9d0a4")
            for tup in users_uuids:
                if tup[0] == variable.get():
                    entry_uuid_auth.insert(0, tup[1])
                    entry_uuid_auth.config(state=DISABLED)
                    entry_uuid_auth.pack()

        point_label = Label(sub_auth_root, text='Выберите точку прохода', font='Times 14', width=28, height=2,
                            bg="#c9d0a4")
        point_label.pack()

        door = IntVar(sub_auth_root)
        door.set(1)

        door_one_radio = Radiobutton(sub_auth_root, text=points[0], variable=door, value=1, font='Times 14', width=28,
                                     height=2, bg="#c9d0a4")
        door_one_radio.pack()

        door_two_radio = Radiobutton(sub_auth_root, text=points[1], variable=door, value=2, font='Times 14', width=28,
                                     height=2, bg="#c9d0a4")
        door_two_radio.pack()

        enter_label = Label(sub_auth_root, text='Выберите тип прохода', font='Times 14', width=28, height=2,
                            bg="#c9d0a4")
        enter_label.pack()

        enter = StringVar(sub_auth_root)
        enter.set('Вход ')

        enter_one_radio = Radiobutton(sub_auth_root, text='Вход', variable=enter, value='Вход', font='Times 14',
                                      width=28, height=2, bg="#c9d0a4")
        enter_one_radio.pack()

        enter_two_radio = Radiobutton(sub_auth_root, text='Выход', variable=enter, value='Выход', font='Times 14',
                                      width=28, height=2, bg="#c9d0a4")
        enter_two_radio.pack()

        def ok():
            info = []
            c = myconnutils.get_connection()
            sql = "Select * from diploma.employee;"
            try:
                curs = c.cursor()
                curs.execute(sql)
                for user_row in curs:
                    info.append(user_row)
            finally:
                # Закрыть соединение (Close connection).
                c.close()

            current_employee = ''

            for employee in info:
                if entry_uuid_auth.get() in employee:
                    current_employee = employee
                    break
                else:
                    current_employee = (
                        'Неизвестно', 'Неизвестно', 'Неизвестно', 'Неизвестно', 'Неизвестно',
                        entry_uuid_auth.get(), 'Неизвестно', 'Неизвестно', 'Неизвестно', 0)

            # Вход
            if current_employee[9] + 1 > 1 and current_employee[0] != 'Неизвестно' and enter.get() == 'Вход':
                text = 'Пользователь с данным пропуском уже вошел!'
                messagebox.showerror("Доступ отказан!", message=text)
                sub_auth_root.destroy()
            elif variable.get() != '-----' and current_employee[9] + 1 < 2 and current_employee[0] != 'Неизвестно' \
                    and enter.get() == 'Вход':
                db_conn = myconnutils.get_connection()
                sql_entry_check = "update diploma.employee set entry_check = entry_check + 1 " \
                                  "where employee.fio='{0}';".format(variable.get())
                try:
                    db_curs = db_conn.cursor()
                    db_curs.execute(sql_entry_check)
                    db_conn.commit()
                finally:
                    # Закрыть соединение (Close connection).
                    db_conn.close()

            ##########################
            # Выход
            if enter.get() == 'Выход' and current_employee[0] != 'Неизвестно' and current_employee[9] - 1 < 0:
                text = 'Пользователь с данным пропуском уже вышел!'
                messagebox.showerror("Доступ отказан!", message=text)
                sub_auth_root.destroy()
            elif enter.get() == 'Выход' and variable.get() != '-----' and current_employee[9] - 1 == 0:
                db_conn = myconnutils.get_connection()
                sql_entry_check = "update diploma.employee set entry_check = entry_check - 1 " \
                                  "where employee.fio='{0}';".format(variable.get())
                try:
                    db_curs = db_conn.cursor()
                    db_curs.execute(sql_entry_check)
                    db_conn.commit()
                finally:
                    # Закрыть соединение (Close connection).
                    db_conn.close()

            ##########################

            if entry_uuid_auth.get() in uuids and sub_entry.get() != '-----':
                c = myconnutils.get_connection()
                sql_ins = "Insert into diploma.log(entry_type,employee_id, uuid ,dep_id, point_id, reg_date)" \
                          " values ('{0}',{1},'{2}',{3},{4},now());".format(enter.get(), current_employee[0],
                                                                            entry_uuid_auth.get(),
                                                                            current_employee[6], door.get())
                try:
                    curs = c.cursor()
                    curs.execute(sql_ins)
                    c.commit()
                    if enter.get() == 'Вход':
                        if datetime.datetime.now().time().hour > 8:

                            co = myconnutils.get_connection()
                            sql_late = "UPDATE diploma.employee SET late_odd = late_odd - 0.05 " \
                                       "WHERE employee.id_employee = {0};".format(current_employee[0])
                            try:
                                pass
                                curs = co.cursor()
                                curs.execute(sql_late)
                                co.commit()
                            finally:
                                # Закрыть соединение (Close connection).
                                co.close()

                            delta_hour = datetime.datetime.now().time().hour - 8
                            delta_min = datetime.datetime.now().time().minute - 0
                            text = 'Доступ разрешен, сотрудник, добро пожаловать! Время прохода {0},' \
                                   ' вы опоздали на {1} часов, {2} минут,' \
                                   ' это негативно отразится на вашей заработной плате!'.format(
                                datetime.datetime.now().time(),
                                delta_hour, delta_min)
                        else:
                            co = myconnutils.get_connection()
                            sql_late = "UPDATE diploma.employee SET late_odd = late_odd + 0.02 " \
                                       "WHERE employee.id_employee = {0};".format(current_employee[0])
                            try:
                                curs = co.cursor()
                                curs.execute(sql_late)
                                co.commit()
                            finally:
                                # Закрыть соединение (Close connection).
                                co.close()
                            text = 'Доступ разрешен, сотрудник, вы пришли вовремя, удачного рабочего дня!'
                    else:
                        text = 'Доступ разрешен, сотрудник, досвидания! Время прохода {0}'.format(
                            datetime.datetime.now().time())
                    messagebox.showinfo("Доступ разрешен!", message=text)
                    sub_auth_root.destroy()

                finally:
                    # Закрыть соединение (Close connection).
                    c.close()

            elif entry_uuid_auth.get() in uuids and sub_entry.get() == '-----':
                connec = myconnutils.get_connection()
                sql_ins = "Insert into diploma.log(entry_type,employee_id, uuid ,dep_id, point_id, reg_date)" \
                          " values ('{0}',{1},'{2}',{3},{4},now());".format(enter.get(), 0, entry_uuid_auth.get(),
                                                                            0, door.get())
                try:
                    curs = connec.cursor()
                    curs.execute(sql_ins)
                    connec.commit()
                    if enter.get() == 'Вход':
                        text = 'Доступ разрешен, гость, добро пожаловать! Время прохода {0}'.format(
                            datetime.datetime.now().time())
                    else:
                        text = 'Доступ разрешен, гость, досвидания! Время прохода {0}'.format(
                            datetime.datetime.now().time())
                    messagebox.showinfo("Доступ разрешен!", message=text)
                    sub_auth_root.destroy()
                finally:
                    # Закрыть соединение (Close connection).
                    connec.close()
            else:
                text = 'Пропуск не действителен, оставайтесь на месте до прихода администратора!'
                messagebox.showerror("Доступ отказан!", message=text)
                error_text = int(door.get())
                email_send_back(error_text)
                sub_auth_root.destroy()

        button = Button(sub_auth_root, text="OK", command=ok, font='Times 14', width=28, height=2, bg='#e8ddaf',
                        foreground='#817162')
        button.pack()

    btn_submit = Button(sub_auth_root, text='Submit', command=submit, font='Times 14', width=28, height=2, bg='#e8ddaf',
                        foreground='#817162')
    btn_submit.pack()

    # test stuff

    sub_auth_root.mainloop()


root = Tk()
root.title("Безопасность будущего")
root.geometry("350x590")
root["bg"] = "#c9d0a4"

btn_auth = Button(text='Аутентификация пользователя', command=user_auth, font='Times 14', width=28, height=2,
                  bg='#e8ddaf', foreground='#817162')
btn_auth.place(x=30, y=20)

btn_create = Button(text='Создание нового пользователя', command=user_create, font='Times 14', width=28, height=2,
                    bg='#e8ddaf', foreground='#817162')
btn_create.place(x=30, y=90)

btn_update = Button(text='Изменение данных пользователя', command=user_update, font='Times 14', width=28, height=2,
                    bg='#e8ddaf', foreground='#817162')
btn_update.place(x=30, y=160)

btn_send = Button(text='Информация о пользователе', command=show_info, font='Times 14', width=28, height=2,
                  bg='#e8ddaf', foreground='#817162')
btn_send.place(x=30, y=230)

btn_all_users = Button(text='Информация о всех пользователях', command=show_all_info, font='Times 14', width=28,
                       height=2, bg='#e8ddaf', foreground='#817162')
btn_all_users.place(x=30, y=300)

btn_show_log = Button(text='Просмотр журнала событий', command=btn_show_log, font='Times 14', width=28, height=2,
                      bg='#e8ddaf', foreground='#817162')
btn_show_log.place(x=30, y=370)

btn_show_graphic_1 = Button(text='График посещения', command=show_graphic_1, font='Times 14', width=28, height=2,
                            bg='#e8ddaf', foreground='#817162')
btn_show_graphic_1.place(x=30, y=440)

btn_show_diagram_2 = Button(text='Диаграмма использования точек', command=show_diagram_2, font='Times 14', width=28,
                            height=2, bg='#e8ddaf', foreground='#817162')
btn_show_diagram_2.place(x=30, y=510)

root.mainloop()
